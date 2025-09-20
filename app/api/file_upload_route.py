import re
import logging
import requests
from flask import request
from flask_restful import Resource
from ..config import Config
from ..services.file_service import FileService
from ..services.solr_service import SolrService
from ..services.database_service import DatabaseService
from openpyxl import load_workbook
from io import BytesIO
import datetime
import threading
import os
from flask import send_file
from ..utils import Utils
import pysolr

logger = logging.getLogger(__name__)




class SingleFileUpload(Resource):
    def __init__(self):
        self.solr_service = SolrService()
        self.db_service = DatabaseService()

    def post(self):
        description = request.form.get('description', '')
        research_name = request.form.get('researchName', '')
        author = request.form.get('author', '')
        file = request.files.get('file')

        logger.info(f"Received file upload request for {file.filename} with research name '{research_name}'")
        resources_created = {
            'solr': False,
            'local_file': None,
            'db': None
        }

        if not file:
            logger.error("No file provided in the request")
            return {
                "status": 0,
                "data": None,
                "message": "Không có tệp nào được cung cấp"
            }, 400

        if not file.filename or research_name == '' or description == '':
            logger.error("File provided has no filename")
            return {
                "status": 0,
                "data": None,
                "message": "Nhập thiếu dữ liệu: tên nghiên cứu, mô tả hoặc tên tệp"
            }, 400


        logger.info(f"Processing file: {file.filename}")

        # Tinh SHA1 cua file
        content = file.read()
        sha1_file = FileService.calculate_sha1(content)
        file.seek(0)

        # Check if document already exists in Solr
        check_response = requests.get(f"{Config.SOLR_URL}/query?q=id:{sha1_file}")

        if check_response.status_code == 200:
            result = check_response.json()
            if result.get("response", {}).get("numFound", 0) > 0:
                msg = f"Tài liệu {file.filename} đã tồn tại trong cơ sở dữ liệu"
                logger.info(msg)
                return {
                    "status": 0,
                    "data": None,
                    "message": msg
                }, 400

        try:
            response = self.solr_service.upload_file(
                sha1_file=sha1_file,
                filename=file.filename,
                content=content,
                mimetype=file.mimetype,
                description=description,
                overwrite="false"
            )

            if response.status_code != 200:
                error_msg = f"Không thể kết nối đến máy chủ cho tệp {file.filename}. Status code: {response.status_code}"
                logger.error(error_msg)
                return {
                    "status": 0,
                    "data": None,
                    "message": error_msg
                }, 500

            result = response.json()
            if result.get("responseHeader", {}).get("status", 0) != 0:
                error_msg = f"Không thể gửi tệp {file.filename} đến máy chủ. Solr response status: {result.get('responseHeader', {}).get('status', 'unknown')}"
                logger.error(error_msg)
                return {
                    "status": 0,
                    "data": None,
                    "message": error_msg
                }, 500

            resources_created['solr'] = True

            file_path = FileService.save_original_file(file, sha1_file, file.filename)
            resources_created['local_file'] = file_path

            document = self.db_service.create_document(
                research_name=research_name,
                file_name=file.filename,
                file_hash=sha1_file,
                description=description,
                file_path=file_path,
                mimetype=file.mimetype,
                file_size=len(content),
                author=author,
            )
            resources_created['db'] = document



            logger.info("Successfully committed changes to Solr")

            # Save data to database
            success_msg = f"Đã gửi tệp {file.filename} đến máy chủ thành công"

            return {
                "status": 1,
                "data": file.filename,
                "message": success_msg
            }, 200

        except Exception as e:
            logger.error(e)
            try:
                if resources_created['solr']:
                    if self.solr_service.delete_file(sha1_file):
                        self.solr_service.commit_changes()
                        logger.info(f"Solr entry deleted during rollback for hash: {sha1_file}")
                if resources_created['local_file']:
                    FileService.delete_file(resources_created['local_file'])
                    logger.info(f"Local file deleted during rollback: {resources_created['local_file']}")
                if resources_created['db']:
                    self.db_service.db.session.delete(resources_created['db'])
                    self.db_service.db.session.commit()
                    logger.info("Database record deleted during rollback")

            except Exception as rollback_error:
                logger.error(f"Error deleting file from Solr or rolling back database: {str(rollback_error)}")
                pass

            return {
                "status": 0,
                "data": None,
                "message": str(e)
            }, 500





class DownloadExcelSample(Resource):
    def get(self):
        """
        Download a sample Excel file for file upload
        """
        try:
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            excel_dir = os.path.join(base_dir, 'excel_sample')
            file_path = os.path.join(excel_dir, 'sample_upload.xlsx')

            return send_file(
                file_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='sample_upload.xlsx'
            )

        except Exception as e:
            logger.error(f"Error downloading Excel sample: {str(e)}")
            return {
                "status": 0,
                "data": None,
                "message": "Lỗi khi tải xuống mẫu tệp Excel"
            }, 500

class MultipleFileSearch(Resource):
    def __init__(self):
        self.solr_service = SolrService()
        self.db_service = DatabaseService()
        from .. import create_app
        self.app = create_app()

    def process_single_file(self, file_info, expmin, expmax, multisource, document_id):
        """
        Process a single file for plagiarism detection
        """
        
        with self.app.app_context():

            scan_status_id = None
            try:
                file_name = file_info['file_name']
                file_content = file_info['file_content']
                file_mimetype = file_info['file_mimetype']
                description = file_info.get('description', '')
                logger.info(f"Processing file in thread: {file_name}")

                # Create a scan status record for this document
                scan_status = self.db_service.create_scan_status(
                    document_id=document_id,
                    status='processing'
                )

                scan_status_id = scan_status.id
                
                # Read file content
                content = file_content
                sha1_file = FileService.calculate_sha1(content)
                description = file_info.get('description', '')

                logger.info(f"Uploading file {file_name} to Solr")


                is_in_solr = False
                check_response = requests.get(f"{Config.SOLR_URL}/query?q=id:{sha1_file}")
                if check_response.status_code == 200:
                    result = check_response.json()
                    if result.get("response", {}).get("numFound", 0) > 0:
                        is_in_solr = True
                        logger.info(f"Document {file_name} already exists in Solr with hash: {sha1_file}")

                if not is_in_solr:
                    upload_response = self.solr_service.upload_file(
                        sha1_file=sha1_file,
                        filename=file_name,
                        content=content,
                        mimetype=file_mimetype,
                        description=description,
                        overwrite="false"
                    )

                    if upload_response.status_code != 200:
                        error_msg = f"Failed to upload file to Solr: {file_name}. Status code: {upload_response.status_code}"
                        logger.error(error_msg)
                        self.db_service.update_scan_status(scan_status_id=scan_status_id, status='failed')
                        return

                    upload_result = upload_response.json()
                    if upload_result.get("responseHeader", {}).get("status", 0) != 0:
                        error_msg = f"Failed to upload file to Solr: {file_name}. Solr response status: {upload_result.get('responseHeader', {}).get('status', 'unknown')}"
                        logger.error(error_msg)
                        self.db_service.update_scan_status(scan_status_id=scan_status_id, status='failed')
                        return

                    # Commit the changes to make sure the file is available for extraction
                    commit_result = self.solr_service.commit_changes()
                    if not commit_result:
                        error_msg = "Failed to commit changes to Solr"
                        logger.error(error_msg)
                        self.db_service.update_scan_status(scan_status_id=scan_status_id, status='failed')
                        return

                # Extract text using Solr
                response = self.solr_service.extract_text(file_name, content, file_mimetype)

                if response.status_code != 200:
                    logger.error(f"Failed to extract text from {file_name}")
                    self.db_service.update_scan_status(scan_status_id=scan_status_id, status='failed')
                    return

                result = response.json()
                if result.get("responseHeader", {}).get("status", 0) != 0:
                    logger.error(f"Solr extraction error for {file_name}")
                    self.db_service.update_scan_status(scan_status_id=scan_status_id, status='failed')
                    return

                document = result.get('file', "")
                rows = 10 if multisource else 1

                # Initialize metrics for plagiarism detection
                sources = {}
                words_doctotal = len(document.split())
                words_scanned = 0
                words_copied = 0
                chars_doctotal = len(document)
                chars_scanned = 0
                chars_copied = 0
                samples_scanned = 0
                samples_copied = 0
                current_sources = None
                output = []

                # Split document into lines and process for plagiarism
                lines = [line.strip() for line in document.split('\n') if line.strip()]

                for line_num, text in enumerate(lines, 1):
                    while True:
                        match = re.search(r'(\S*\w\S*([\s.,-]+|$)+){' + str(expmin) + ',' + str(expmax) + '}', text)
                        if not match:
                            break
                        sample = match.group(0).strip()
                        possample = match.start()

                        presample = text[:possample]
                        text = text[possample + len(sample):]

                        samples_scanned += 1
                        words_in_sample = len(sample.split())
                        words_scanned += words_in_sample
                        chars_scanned += len(sample)

                        output.append({"type": "text", "content": presample})

                        # Search for matches in Solr
                        data = {
                            "q": f'"{Utils.escape_solr_text(sample)}" AND NOT id:"{sha1_file}"',
                            "fl": "id,resource_name,description",
                            "rows": rows
                        }
                        response = requests.post(f"{Config.SOLR_URL}/query", data=data)
                        result = response.json()

                        if result.get("response", {}).get("numFound", 0) > 0:
                            samples_copied += 1
                            words_copied += words_in_sample
                            chars_copied += len(sample)
                            new_sources = {}

                            for doc in result["response"]["docs"]:
                                source_id = doc["id"]
                                sources[source_id] = sources.get(source_id, {
                                    "color": source_id[:6],
                                    "name": doc.get("resource_name", ["Unknown"])[0],
                                    "description": doc.get("description", [""])[0],
                                    "words": 0,
                                    "samples": 0
                                })
                                sources[source_id]["words"] += words_in_sample
                                sources[source_id]["samples"] += 1
                                new_sources[source_id] = True

                            if current_sources != new_sources:
                                current_sources = new_sources
                                for source_id in new_sources:
                                    output.append({
                                        "type": "marker",
                                        "id": f"{sha1_file}_{source_id}",
                                        "color": sources[source_id]["color"],
                                        "name": sources[source_id]["name"]
                                    })

                            output.append({
                                "type": "highlight",
                                "content": sample
                            })
                        else:
                            output.append({"type": "text", "content": sample})

                    output.append({"type": "text", "content": text})
                    output.append({"type": "br"})

                # Calculate ratios
                chars_original = chars_scanned - chars_copied
                chars_original_ratio = chars_original / chars_scanned if chars_scanned else 0
                words_original = words_scanned - words_copied
                words_original_ratio = words_original / words_scanned if words_scanned else 0
                samples_original = samples_scanned - samples_copied
                samples_original_ratio = samples_original / samples_scanned if samples_scanned else 0

                # Sort sources by words
                sorted_sources = sorted(sources.items(), key=lambda x: x[1]["words"], reverse=True)

                # Create metrics and parameters dictionaries
                metrics = {
                    "chars_doctotal": chars_doctotal,
                    "chars_scanned": chars_scanned,
                    "chars_original": chars_original,
                    "chars_copied": chars_copied,
                    "chars_original_ratio": chars_original_ratio,
                    "words_doctotal": words_doctotal,
                    "words_scanned": words_scanned,
                    "words_original": words_original,
                    "words_copied": words_copied,
                    "words_original_ratio": words_original_ratio,
                    "samples_scanned": samples_scanned,
                    "samples_original": samples_original,
                    "samples_copied": samples_copied,
                    "samples_original_ratio": samples_original_ratio
                }

                parameters = {
                    "exp_min": expmin,
                    "exp_max": expmax,
                    "multi_source": multisource
                }

                # Create result object with output data
                output_data = {
                    "filename": file_name,
                    "output": output
                }

                # Create scan result in database
                scan_result = self.db_service.create_scan_result(
                    status_id=scan_status_id,
                    metrics=metrics,
                    parameters=parameters,
                    output_data=output_data
                )

                # Create scan resources records for sources
                sources_list = [
                    {
                        "id": source_id,
                        "color": info["color"],
                        "name": info["name"],
                        "description": info["description"],
                        "words": info["words"],
                        "samples": info["samples"]
                    } for source_id, info in sorted_sources
                ]

                self.db_service.create_scan_resources(scan_result.id, sources_list)

                # Mark scan as completed
                self.db_service.update_scan_status(
                scan_status_id=scan_status_id,
                status='completed',
                finished_date=datetime.datetime.utcnow()
                )

                logger.info(f"Successfully processed file: {file_name}")

            except Exception as e:
                logger.error(f"Error processing file {file_info['file'].filename}: {str(e)}")
                if 'scan_status_id' in locals():
                    self.db_service.update_scan_status(scan_status_id=scan_status_id, status='failed')
    def post(self):
        files = request.files.getlist('files')
        excel = request.files.get('excel')
        expmin = int(request.form.get('expmin', '3'))
        expmax = int(request.form.get('expmax', '5'))
        multisource = request.form.get('multisource', 'false').lower() == 'true'

        if expmin < 1 or expmax < expmin:
            return {
                "status": 0,
                "data": None,
                "message": "Giá trị expmin hoặc expmax không hợp lệ"
            }, 400

        if not files:
            return {
                "status": 0,
                "data": None,
                "message": "Không có tệp nào được cung cấp"
            }, 400

        if not all(file.filename for file in files):
            return {
                "status": 0,
                "data": None,
                "message": "Một hoặc nhiều tệp không có tên"
            }, 400

        if excel:
            if not excel.filename.endswith(('.xlsx', '.xls')):
                return {
                    "status": 0,
                    "data": None,
                    "message": "Tệp Excel phải có định dạng .xlsx"
                }, 400
        else:
            return {
                "status": 0,
                "data": None,
                "message": "Tệp Excel không được cung cấp"
            }, 400

        try:
            file_stream = BytesIO(excel.read())
            workbook = load_workbook(file_stream)
            sheet = workbook.active

            search_data = []

            for file in files:
                file_matched = False
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1] and row[0] == file.filename:
                        # Found a match in the Excel file
                        file_info = {
                            'file': file,
                            'research_name': row[1],
                            'description': row[2] if len(row) > 2 and row[2] else ''
                        }
                        search_data.append(file_info)
                        file_matched = True
                        logger.info(f"Matched file {file.filename} with Excel data")
                        break

                if not file_matched:
                    logger.warning(f"No matching entry in Excel for file: {file.filename}")

            if not search_data:
                return {
                    "status": 0,
                    "data": None,
                    "message": "Không có tệp nào khớp với dữ liệu trong tệp Excel"
                }, 400

            logger.info(f"Successfully matched {len(search_data)} files with Excel data")

            # Save each file to DB with pending status and create documents
            document_ids = []
            new_documents = []
            existing_documents = []


            for file_info in search_data:
                file = file_info['file']
                content = file.read()
                file_info['file_content'] = content
                file_info['file_name'] = file.filename
                file_info['file_mimetype'] = file.mimetype
                sha1_file = FileService.calculate_sha1(file_info['file_content'])
                file.seek(0)
                

                existing_document = self.db_service.get_document_by_hash(sha1_file)
                if existing_document:
                    # Document already exists, use its ID
                    document_ids.append(existing_document.id)
                    existing_documents.append({
                        'id': existing_document.id,
                        'filename': file_info['file_name'],
                        'hash': sha1_file
                    })
                    logger.info(f"Document {file_info['file_name']} already exists in database with ID: {existing_document.id}")
                else:
                    # Save the original file to disk
                    file_path = FileService.save_original_file(file, sha1_file, file_info['file_name'])
                    document = self.db_service.create_document(
                        research_name=file_info['research_name'],
                        file_name=file_info['file_name'],
                        file_hash=sha1_file,
                        description=file_info['description'],
                        file_path=file_path,
                        mimetype=file_info['file_mimetype'],
                        file_size=len(file_info['file_content'])
                    )
                    document_ids.append(document.id)
                    new_documents.append({
                        'id': document.id,
                        'filename': file_info['file_name'],
                        'hash': sha1_file
                    })
                    logger.info(f"Created new document for {file_info['file_name']} with ID: {document.id}")
                
                file.close()  # Close the file stream after saving
            
            # Process each file in separate threads sequentially
            threads = []
            for i, file_info in enumerate(search_data):
                thread = threading.Thread(
                    target=self.process_single_file,
                    args=(file_info, expmin, expmax, multisource, document_ids[i])
                )
                threads.append(thread)

            # Create a thread coordinator to run threads sequentially
            def run_threads_sequentially():
                for i, thread in enumerate(threads):
                    logger.info(f"Starting thread for processing file: {search_data[i]['file'].filename}")
                    thread.start()
                    thread.join()  # Wait for thread to complete before starting the next one
                logger.info("All file processing threads completed")

            # Start the coordinator thread
            coordinator = threading.Thread(target=run_threads_sequentially)
            coordinator.daemon = True
            coordinator.start()

            return {
                "status": 1,
                "data": {
                    "document_ids": document_ids,
                    "new_documents": new_documents,
                    "existing_documents": existing_documents,
                    "matched_files": [item['file'].filename for item in search_data]
                },
                "message": f"Đã tìm thấy {len(search_data)} tệp và bắt đầu phân tích. Vui lòng kiểm tra trạng thái sau."
            }, 200

        except Exception as e:
            logger.error(f"Error setting up file processing: {str(e)}")
            return {
                "status": 0,
                "data": None,
                "message": f"Lỗi khi xử lý: {str(e)}"
            }, 500

# class SingleFileSearch(Resource):
#     def __init__(self):
#         self.solr_service = SolrService()
#
#     def post(self):
#
#         expmin = int(request.form.get('expmin', '3'))
#         expmax = int(request.form.get('expmax', '5'))
#         multisource = request.form.get('multisource', 'false').lower() == 'true'
#         file = request.files.get('file')
#
#         if expmin < 1 or expmax < expmin:
#             return {
#                 "status": 0,
#                 "data": None,
#                 "message": "Giá trị expmin hoặc expmax không hợp lệ"
#             }, 400
#
#         if not file:
#             return {
#                 "status": 0,
#                 "data": None,
#                 "message": "Không có tệp nào được cung cấp"
#             }, 400
#         if not file.filename:
#             return {
#                 "status": 0,
#                 "data": None,
#                 "message": "Tệp không có tên"
#             }, 400
#
#         try:
#             content = file.read()
#             sha1_file = FileService.calculate_sha1(content)
#             file.seek(0)
#
#             response = self.solr_service.extract_text(
#                 filename=file.filename,
#                 content=content,
#                 mimetype=file.mimetype
#             )
#
#             if response.status_code != 200:
#                 error_msg = f"Không thể kết nối đến máy chủ Solr cho tệp {file.filename}. Status code: {response.status_code}"
#                 logger.error(error_msg)
#                 return {
#                     "status": 0,
#                     "data": None,
#                     "message": error_msg
#                 }, 500
#
#             result = response.json()
#
#             if result.get("responseHeader", {}).get("status", 0) != 0:
#                 error_msg = f"Không thể trích xuất văn bản từ {file.filename}"
#                 logger.error(
#                     f"{error_msg}. Solr response status: {result.get('responseHeader', {}).get('status', 'unknown')}")
#                 return {
#                     "status": 0,
#                     "data": None,
#                     "message": error_msg
#                 }, 500
#
#             document = result.get('file', "")
#
#             if file.mimetype == 'application/pdf' or file.filename.lower().endswith('.pdf'):
#                 document = self._clean_pdf_text(document)
#
#             rows = 10 if multisource else 1
#
#             # Initialize metrics
#             sources = {}
#             words_doctotal = len(document.split())
#             words_scanned = 0
#             words_copied = 0
#             chars_doctotal = len(document)
#             chars_scanned = 0
#             chars_copied = 0
#             samples_scanned = 0
#             samples_copied = 0
#             current_sources = None
#             output = []
#
#             #Split document into lines
#             lines = [line.strip() for line in document.split('\n') if line.strip()]
#
#             for line_num, text in enumerate(lines, 1):
#                 while True:
#                     match = re.search(r'(\S*\w\S*([\s.,-]+|$)+){' + str(expmin) + ',' + str(expmax) + '}', text)
#                     if not match:
#                         break
#                     sample, possample = match.group(0), match.start()
#                     presample = text[:possample]
#                     text = text[possample + len(sample):]
#
#                     sample = self._clean_search_sample(sample)
#                     if not sample or len(sample.strip()) < 3:  # Skip very short samples
#                         output.append({"type": "text", "content": presample + sample})
#                         continue
#
#
#
#                     samples_scanned += 1
#                     words_in_sample = len(sample.split())
#                     words_scanned += words_in_sample
#                     chars_scanned += len(sample)
#
#                     output.append({"type": "text", "content": presample})
#
#                     # Search for sample in Solr
#                     logger.debug(f"Searching for sample {samples_scanned} in Solr database")
#                     data = {
#                         "q": f'"{Utils.escape_solr_text(sample)}"',
#                         "fl": "id,resource_name,description",
#                         "rows": rows
#                     }
#                     response = requests.post(f"{Config.SOLR_URL}/query", data=data)
#                     result = response.json()
#
#                     if result.get("response", {}).get("numFound", 0) > 0:
#                         samples_copied += 1
#                         words_copied += words_in_sample
#                         chars_copied += len(sample)
#                         new_sources = {}
#
#                         logger.debug(f"Found {result['response']['numFound']} matches for sample {samples_scanned}")
#
#                         for doc in result["response"]["docs"]:
#                             source_id = doc["id"]
#                             sources[source_id] = sources.get(source_id, {
#                                 "color": source_id[:6],
#                                 "name": doc.get("resource_name", ["Unknown"])[0],
#                                 "description": doc.get("description", [""])[0],
#                                 "words": 0,
#                                 "samples": 0
#                             })
#                             sources[source_id]["words"] += words_in_sample
#                             sources[source_id]["samples"] += 1
#                             new_sources[source_id] = True
#
#                         if current_sources != new_sources:
#                             current_sources = new_sources
#                             for source_id in new_sources:
#                                 output.append({
#                                     "type": "marker",
#                                     "id": f"{sha1_file}_{source_id}",
#                                     "color": sources[source_id]["color"],
#                                     "name": sources[source_id]["name"]
#                                 })
#
#                         output.append({
#                             "type": "highlight",
#                             "content": sample
#                         })
#                     else:
#                         logger.debug(f"No matches found for sample {samples_scanned}")
#                         output.append({"type": "text", "content": sample})
#
#                 output.append({"type": "text", "content": text})
#                 output.append({"type": "br"})
#
#             # Calculate ratios
#             chars_original = chars_scanned - chars_copied
#             chars_original_ratio = chars_original / chars_scanned if chars_scanned else 0
#             words_original = words_scanned - words_copied
#             words_original_ratio = words_original / words_scanned if words_scanned else 0
#             samples_original = samples_scanned - samples_copied
#             samples_original_ratio = samples_original / samples_scanned if samples_scanned else 0
#
#             logger.info(f"Plagiarism analysis completed for {file.filename}")
#             logger.info(f"Samples scanned: {samples_scanned}, copied: {samples_copied}, original: {samples_original}")
#             logger.info(f"Words scanned: {words_scanned}, copied: {words_copied}, original: {words_original}")
#             logger.info(f"Characters scanned: {chars_scanned}, copied: {chars_copied}, original: {chars_original}")
#             logger.info(
#                 f"Originality ratio - Words: {words_original_ratio:.2%}, Characters: {chars_original_ratio:.2%}")
#
#             # Sort sources by words
#             sorted_sources = sorted(sources.items(), key=lambda x: x[1]["words"], reverse=True)
#             logger.debug(f"Found {len(sorted_sources)} sources for plagiarism matches")
#
#             result = {
#                 "filename": file.filename,
#                 "metrics": {
#                     "chars_doctotal": chars_doctotal,
#                     "chars_scanned": chars_scanned,
#                     "chars_original": chars_original,
#                     "chars_copied": chars_copied,
#                     "chars_original_ratio": chars_original_ratio,
#                     "words_doctotal": words_doctotal,
#                     "words_scanned": words_scanned,
#                     "words_original": words_original,
#                     "words_copied": words_copied,
#                     "words_original_ratio": words_original_ratio,
#                     "samples_scanned": samples_scanned,
#                     "samples_original": samples_original,
#                     "samples_copied": samples_copied,
#                     "samples_original_ratio": samples_original_ratio
#                 },
#                 "sources": [
#                     {
#                         "id": source_id,
#                         "color": info["color"],
#                         "name": info["name"],
#                         "description": info["description"],
#                         "words": info["words"],
#                         "samples": info["samples"]
#                     } for source_id, info in sorted_sources
#                 ],
#                 "output": output
#             }
#
#             return {
#                 "status": 1,
#                 "data": result,
#                 "message": "Phân tích đạo văn thành công"
#             }, 200
#
#         except Exception as e:
#             logger.error(str(e))
#             return {
#                 "status": 0,
#                 "data": None,
#                 "message": "Lỗi hệ thống: " + str(e)
#             }, 500
#
#     def _clean_pdf_text(self, text):
#         """Clean PDF extracted text from common artifacts"""
#         if not text:
#             return text
#
#         # Remove common PDF artifacts
#         text = re.sub(r'\x00', '', text)  # Remove null bytes
#         text = re.sub(r'[\x01-\x08\x0B\x0C\x0E-\x1F\x7F]', '', text)  # Remove control characters
#         text = re.sub(r'\ufeff', '', text)  # Remove BOM
#         text = re.sub(r'[\u200b-\u200f\u2028-\u202f\u205f-\u206f]', ' ', text)  # Remove zero-width chars
#
#         # Normalize whitespace
#         text = re.sub(r'\s+', ' ', text)
#         text = text.strip()
#
#         return text
#
#     def _clean_search_sample(self, sample):
#         """Clean and validate search sample"""
#         if not sample:
#             return sample
#
#         # Remove problematic characters that can break Solr queries
#         sample = re.sub(r'[^\w\s\.,;:!?\-\'\"()]', ' ', sample)  # Keep only safe characters
#         sample = re.sub(r'\s+', ' ', sample)  # Normalize whitespace
#         sample = sample.strip()
#
#         # Remove very short or very long samples
#         if len(sample) < 3 or len(sample) > 1000:
#             return ""
#
#         return sample

class SingleFileSearch(Resource):
    def __init__(self):
        self.solr_service = SolrService()

    def process_document_optimized(self, document, sha1_file, expmin, expmax, multisource):
        """Optimized document processing with individual accurate searches"""

        samples_with_positions = []
        lines = [line.strip() for line in document.split('\n') if line.strip()]

        # Extract all samples with proper positioning
        for line_num, text in enumerate(lines, 1):
            line_start = 0
            while True:
                match = re.search(r'(\S*\w\S*([\s.,-]+|$)+){' + str(expmin) + ',' + str(expmax) + '}', text[line_start:])
                if not match:
                    break

                sample = match.group(0).strip()
                sample = self._clean_search_sample(sample)

                if sample and len(sample.strip()) >= 3:
                    samples_with_positions.append({
                        'index': len(samples_with_positions),
                        'sample': sample,
                        'line_num': line_num,
                        'start_pos': line_start + match.start(),
                        'end_pos': line_start + match.end(),
                        'text_context': text
                    })

                line_start += match.end()

        logger.info(f"Found {len(samples_with_positions)} samples to process")

        # Use individual searches with connection reuse instead of batch
        samples_for_search = [(item['index'], item['sample']) for item in samples_with_positions]
        search_results = self.solr_service.search_samples_individual(samples_for_search, sha1_file)

        logger.info(f"Completed individual searches, found matches for {len(search_results)} samples")
        return self._build_output_with_results(document, samples_with_positions, search_results, sha1_file, multisource)

    def _build_output_with_results(self, document, samples_with_positions, search_results, sha1_file, multisource):
        """Build output using pre-computed search results"""

        sources = {}
        words_doctotal = len(document.split())
        words_scanned = 0
        words_copied = 0
        chars_doctotal = len(document)
        chars_scanned = 0
        chars_copied = 0
        samples_scanned = 0
        samples_copied = 0
        current_sources = None
        output = []

        lines = [line.strip() for line in document.split('\n') if line.strip()]

        for line_num, text in enumerate(lines, 1):
            line_samples = [s for s in samples_with_positions if s['line_num'] == line_num]
            line_samples.sort(key=lambda x: x['start_pos'])

            current_pos = 0

            for sample_info in line_samples:
                sample = sample_info['sample']
                start_pos = sample_info['start_pos']
                end_pos = sample_info['end_pos']
                sample_idx = sample_info['index']

                # Add text before sample
                presample = text[current_pos:start_pos]
                output.append({"type": "text", "content": presample})

                # Update metrics
                samples_scanned += 1
                words_in_sample = len(sample.split())
                words_scanned += words_in_sample
                chars_scanned += len(sample)

                # Check if sample has matches
                if sample_idx in search_results:
                    samples_copied += 1
                    words_copied += words_in_sample
                    chars_copied += len(sample)
                    new_sources = {}

                    docs = search_results[sample_idx][:10 if multisource else 1]

                    for doc in docs:
                        source_id = doc["id"]
                        sources[source_id] = sources.get(source_id, {
                            "color": source_id[:6],
                            "name": doc.get("resource_name", "Unknown"),
                            "description": doc.get("description", ""),
                            "words": 0,
                            "samples": 0
                        })
                        sources[source_id]["words"] += words_in_sample
                        sources[source_id]["samples"] += 1
                        new_sources[source_id] = True

                    if current_sources != new_sources:
                        current_sources = new_sources
                        for source_id in new_sources:
                            output.append({
                                "type": "marker",
                                "id": f"{sha1_file}_{source_id}",
                                "color": sources[source_id]["color"],
                                "name": sources[source_id]["name"]
                            })

                    output.append({"type": "highlight", "content": sample})
                else:
                    output.append({"type": "text", "content": sample})

                current_pos = end_pos

            # Add remaining text in line
            remaining_text = text[current_pos:]
            output.append({"type": "text", "content": remaining_text})
            output.append({"type": "br"})

        # Calculate ratios
        chars_original = chars_scanned - chars_copied
        chars_original_ratio = chars_original / chars_scanned if chars_scanned else 0
        words_original = words_scanned - words_copied
        words_original_ratio = words_original / words_scanned if words_scanned else 0
        samples_original = samples_scanned - samples_copied
        samples_original_ratio = samples_original / samples_scanned if samples_scanned else 0

        # Sort sources by words
        sorted_sources = sorted(sources.items(), key=lambda x: x[1]["words"], reverse=True)

        return {
            "filename": "processed_document",
            "metrics": {
                "chars_doctotal": chars_doctotal,
                "chars_scanned": chars_scanned,
                "chars_original": chars_original,
                "chars_copied": chars_copied,
                "chars_original_ratio": chars_original_ratio,
                "words_doctotal": words_doctotal,
                "words_scanned": words_scanned,
                "words_original": words_original,
                "words_copied": words_copied,
                "words_original_ratio": words_original_ratio,
                "samples_scanned": samples_scanned,
                "samples_original": samples_original,
                "samples_copied": samples_copied,
                "samples_original_ratio": samples_original_ratio
            },
            "sources": [
                {
                    "id": source_id,
                    "color": info["color"],
                    "name": info["name"],
                    "description": info["description"],
                    "words": info["words"],
                    "samples": info["samples"]
                } for source_id, info in sorted_sources
            ],
            "output": output
        }

    def post(self):

        expmin = int(request.form.get('expmin', '3'))
        expmax = int(request.form.get('expmax', '5'))
        multisource = request.form.get('multisource', 'false').lower() == 'true'
        file = request.files.get('file')

        if expmin < 1 or expmax < expmin:
            return {
                "status": 0,
                "data": None,
                "message": "Giá trị expmin hoặc expmax không hợp lệ"
            }, 400

        if not file:
            return {
                "status": 0,
                "data": None,
                "message": "Không có tệp nào được cung cấp"
            }, 400
        if not file.filename:
            return {
                "status": 0,
                "data": None,
                "message": "Tệp không có tên"
            }, 400

        try:
            content = file.read()
            sha1_file = FileService.calculate_sha1(content)
            file.seek(0)

            response = self.solr_service.extract_text(
                filename=file.filename,
                content=content,
                mimetype=file.mimetype
            )

            if response.status_code != 200:
                error_msg = f"Không thể kết nối đến máy chủ Solr cho tệp {file.filename}. Status code: {response.status_code}"
                logger.error(error_msg)
                return {
                    "status": 0,
                    "data": None,
                    "message": error_msg
                }, 500

            result = response.json()

            if result.get("responseHeader", {}).get("status", 0) != 0:
                error_msg = f"Không thể trích xuất văn bản từ {file.filename}"
                logger.error(
                    f"{error_msg}. Solr response status: {result.get('responseHeader', {}).get('status', 'unknown')}")
                return {
                    "status": 0,
                    "data": None,
                    "message": error_msg
                }, 500

            document = result.get('file', "")

            if file.mimetype == 'application/pdf' or file.filename.lower().endswith('.pdf'):
                document = self._clean_pdf_text(document)

            result = self.process_document_optimized(document, sha1_file, expmin, expmax, multisource)
            result["filename"] = file.filename

            return {
                "status": 1,
                "data": result,
                "message": "Phân tích đạo văn thành công"
            }, 200


        except Exception as e:
            logger.error(str(e))
            return {
                "status": 0,
                "data": None,
                "message": "Lỗi hệ thống: " + str(e)
            }, 500

    def _clean_pdf_text(self, text):
        """Clean PDF extracted text from common artifacts"""
        if not text:
            return text

        # Remove common PDF artifacts
        text = re.sub(r'\x00', '', text)  # Remove null bytes
        text = re.sub(r'[\x01-\x08\x0B\x0C\x0E-\x1F\x7F]', '', text)  # Remove control characters
        text = re.sub(r'\ufeff', '', text)  # Remove BOM
        text = re.sub(r'[\u200b-\u200f\u2028-\u202f\u205f-\u206f]', ' ', text)  # Remove zero-width chars

        # Normalize whitespace
        text = re.sub(r'\s+', ' ', text)
        text = text.strip()

        return text

    def _clean_search_sample(self, sample):
        """Clean and validate search sample"""
        if not sample:
            return sample

        # Remove problematic characters that can break Solr queries
        sample = re.sub(r'[^\w\s\.,;:!?\-\'\"()]', ' ', sample)  # Keep only safe characters
        sample = re.sub(r'\s+', ' ', sample)  # Normalize whitespace
        sample = sample.strip()

        # Remove very short or very long samples
        if len(sample) < 3 or len(sample) > 1000:
            return ""

        return sample